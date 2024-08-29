import httpx
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.chart import BarChart, Reference
# Path to the xlsx file containing queries
query_file_path = './eval_queries.xlsx'  # Ensure this path is correct

# Load the workbook and select the sheet with queries
query_workbook = load_workbook(filename=query_file_path)
query_sheet = query_workbook['Sheet2']  # Replace 'Sheet2' with your actual sheet name if different

# Create a new workbook for the results
result_workbook = Workbook()
result_sheet = result_workbook.active
result_sheet.title = "Query Results"

# Add headers to the result sheet
result_sheet.append(["Query", "EID", "Chunk Title", "Chunk Text", "Length"])

# Define the URL
url = "https://rag-retrieve-pr-299.dev.knowledge.healthcare.elsevier.systems/retrieve_vector"

# Define the headers
headers = {
    "Content-Type": "application/json; charset=utf-8"
}

# Define the static part of the data
static_data = {
    "dsl_filter": {
        "terms": {
            "source_id": [
                "375706",
                "797588",
                "797586",
                "797505",
                "797506",
                "799812",
                "797589",
                "799811",
                "298363"
            ]
        }
    }
}

# Red fill for cells with chunk_text less than 20 words
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")



# Process each query sequentially
queries = [row[0].value for row in query_sheet.iter_rows(min_col=1, max_col=1, min_row=2) if row[0].value]
word_counts = []
try:
    for query in queries:
        static_data["query_text"] = query
        try:
            response = httpx.post(url, headers=headers, json=static_data, timeout=10.0)
            response.raise_for_status()
            results = response.json()["results"]
            result_sheet.append([query, "", "", "", ""])
            for chunk in results:
                chunk_id = chunk["_source"]["eid"]
                chunk_title = chunk["_source"]["title"]
                chunk_text = chunk["_source"]["chunk_text"]
                chunk_length = len(chunk_text.split())
                word_counts.append(chunk_length)
                result_sheet.append(["", chunk_id, chunk_title, chunk_text, chunk_length])
                if chunk_length <= 20:
                    if chunk_length <= 20:
                        for col in range(2, 6):
                            result_sheet.cell(row=result_sheet.max_row, column=col).fill = red_fill
        except Exception as e:
            print(f"An error occurred while processing query '{query}': {e}")    
        print(f"Processed query: {query}")

except Exception as e:
    print(f"An error occurred: {e}")
finally:

    # Define bins for word count ranges
    bins = [0, 10, 20, 50, 100, float('inf')]
    bin_labels = ["0-10", "10-20", "20-50", "50-100", "100+"]

    # Count the number of chunks in each bin
    bin_counts = [0] * (len(bins) - 1)
    for count in word_counts:
        for i in range(len(bins) - 1):
            if bins[i] <= count < bins[i + 1]:
                bin_counts[i] += 1
                break

    # Add a new sheet for word count distribution
    distribution_sheet = result_workbook.create_sheet(title="Word Count Distribution")

    # Insert word count distribution into the new sheet
    distribution_sheet.append(["Range", "Count"])
    for label, count in zip(bin_labels, bin_counts):
        distribution_sheet.append([label, count])

    # Create a bar chart
    chart = BarChart()
    chart.title = "Distribution of Word Counts in Chunk Texts"
    chart.x_axis.title = "Word Count Range"
    chart.y_axis.title = "Frequency"

    # Define the data for the chart
    data = Reference(distribution_sheet, min_col=2, min_row=1, max_row=len(bin_counts) + 1, max_col=2)
    categories = Reference(distribution_sheet, min_col=1, min_row=2, max_row=len(bin_counts) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    # Add the chart to the sheet
    distribution_sheet.add_chart(chart, "E5")

    # Save the new Excel file with the results
    result_file_path = './query_results.xlsx'
    result_workbook.save(filename=result_file_path)
    query_workbook.close()
    result_workbook.close()