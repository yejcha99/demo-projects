import httpx
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.chart import BarChart, Reference

# Path to the xlsx file
file_path = './eval_queries.xlsx'  # Ensure this path is correct

# Load the workbook and select the sheet
workbook = load_workbook(filename=file_path)
sheet = workbook['Sheet2']  # Replace 'Sheet1' with your actual sheet name if different

# Specify the column to read (e.g., 'A' for the first column)
column_letter = 'A'

# Define the URL
url = "https://rag-retrieve-pr-299.dev.knowledge.healthcare.elsevier.systems/retrieve_vector"

# Define the headers
headers = {
    "Content-Type": "application/json; charset=utf-8"
}

# Define the data
static_data = {
    "dsl_filter": {
        "terms": {
            "source_id": [
                "375706",
                "797588",
                "797586",
                "797505",
                "797506",
                "799812",
                "797589",
                "799811",
                "298363"
            ]
        }
    }
}

# Red fill for cells with chunk_text less than or equal to 20 words
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# List to store word counts of all chunk_texts
word_counts = []

# Iterate over the rows in the specified column and make server requests
for row in sheet.iter_rows(min_col=1, max_col=1, min_row=2):  # Assuming the first row is the header
    cell = row[0]
    query_value = cell.value
    if query_value:
        # Construct the data with the dynamic query_text
        data = static_data.copy()
        data["query_text"] = query_value
        
        # Make the HTTP POST request
        response = httpx.post(url, headers=headers, json=data)
        
        # Retrieve the chunk_texts from the response
        chunk_texts = [result["_source"]["chunk_text"] for result in response.json()["results"]]
        
        # Write each chunk_text to a new column in the same row
        for idx, chunk_text in enumerate(chunk_texts, start=2):  # Start from column B (2)
            output_cell = sheet.cell(row=cell.row, column=idx)
            output_cell.value = chunk_text
            
            # Count the number of words in the chunk_text
            word_count = len(chunk_text.split())
            
            # Change the background color of the cell to red if the chunk_text has 20 or fewer words
            if word_count <= 20:
                output_cell.fill = red_fill
            
            # Add the word count to the list
            word_counts.append(word_count)
    print(f"Processed query: {query_value}")

# Save the updated Excel file with a new name
new_file_path = './updated_eval_queries1.xlsx'  # Specify the new file name
workbook.save(filename=new_file_path)

# Calculate histogram data with specified ranges
ranges = {
    "0-10": 0,
    "10-20": 0,
    "20-50": 0,
    "50-100": 0,
    "100+": 0
}

for count in word_counts:
    if count <= 10:
        ranges["0-10"] += 1
    elif count <= 20:
        ranges["10-20"] += 1
    elif count <= 50:
        ranges["20-50"] += 1
    elif count <= 100:
        ranges["50-100"] += 1
    else:
        ranges["100+"] += 1

# Write histogram data to the sheet
histogram_sheet = workbook.create_sheet(title="Histogram Data")
histogram_sheet.append(["Word Count Range", "Frequency"])
for range_label, frequency in ranges.items():
    histogram_sheet.append([range_label, frequency])

# Create a bar chart
chart = BarChart()
data = Reference(histogram_sheet, min_col=2, min_row=1, max_col=2, max_row=len(ranges) + 1)
categories = Reference(histogram_sheet, min_col=1, min_row=2, max_row=len(ranges) + 1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)
chart.title = "Distribution of Word Counts in Chunk Texts"
chart.x_axis.title = "Word Count Range"
chart.y_axis.title = "Frequency"

# Add the chart to the sheet
histogram_sheet.add_chart(chart, "E5")

# Save the updated Excel file with the histogram
workbook.save(filename=new_file_path)