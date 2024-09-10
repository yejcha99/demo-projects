import httpx
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.chart import BarChart, Reference, Series, ScatterChart
from openpyxl.chart.label import DataLabelList
from collections import defaultdict
import numpy as np

# Path to the xlsx file containing queries
query_file_path = './eval_queries.xlsx'  # Ensure this path is correct

# Load the workbook and select the sheet with queries
query_workbook = load_workbook(filename=query_file_path)
query_sheet = query_workbook['Sheet2']  # Replace 'Sheet2' with your actual sheet name if different

# Create a new workbook for the results
result_workbook = Workbook()
result_sheet = result_workbook.active
result_sheet.title = "Query Results"

# Add headers to the result sheet
result_sheet.append(["Query", "EID", "Content Type", "Chunk Title", "Chunk Text", "Length", "Score"])

# Define the URL
url = "https://rag-retrieve-pr-342.dev.knowledge.healthcare.elsevier.systems/retrieve_vector"

# Define the headers
headers = {
    "Content-Type": "application/json; charset=utf-8"
}

# Define the static part of the data
static_data = {
    "dsl_filter": {
        "terms": {
            "source_id": [
                "375706",
                "797588",
                "797586",
                "797505",
                "797506",
                "799812",
                "797589",
                "799811",
                "298363",
                "797491",
                "9780323930383",
                "9780323722193",
                "9780323883054",
            ]
        }
    }
}

# Red fill for cells with chunk_text less than 20 words
red_fill = PatternFill(start_color="FF3632", end_color="FF3632", fill_type="solid")



# Process each query sequentially
queries = [row[0].value for row in query_sheet.iter_rows(min_col=1, max_col=1, min_row=2) if row[0].value]
word_counts = []
chunk_types_less_than_20 = []
chunk_types_all = []
chunk_lengths = []
chunk_scores = []
try:
    for query in queries:
        static_data["query_text"] = query
        try:
            response = httpx.post(url, headers=headers, json=static_data, timeout=10.0)
            response.raise_for_status()
            results = response.json()["results"]
            result_sheet.append([query, "", "", "", "", ""])
            for chunk in results:
                chunk_id = chunk["_source"]["eid"]
                chunk_type = chunk["_source"]["content_type"]
                chunk_title = chunk["_source"]["title"]
                chunk_text = chunk["_source"]["chunk_text"]
                chunk_length = len(chunk_text.split())
                chunk_score = chunk["_score"]
                word_counts.append(chunk_length)
                chunk_types_all.append(chunk_type)
                chunk_lengths.append(chunk_length)
                chunk_scores.append(chunk_score)
                result_sheet.append(["", chunk_id, chunk_type, chunk_title, chunk_text, chunk_length, chunk_score])
                if chunk_length <= 20:
                    chunk_types_less_than_20.append(chunk_type)
                    for col in range(2, 7):
                        result_sheet.cell(row=result_sheet.max_row, column=col).fill = red_fill
        except Exception as e:
            print(f"An error occurred while processing query '{query}': {e}")    
        print(f"Processed query: {query}")

except Exception as e:
    print(f"An error occurred: {e}")
finally:

    # Define bins for word count ranges
    bins = [0, 10, 20, 50, 100, float('inf')]
    bin_labels = ["0-10", "10-20", "20-50", "50-100", "100+"]

    # Count the number of chunks in each bin
    bin_counts = [0] * (len(bins) - 1)
    for count in word_counts:
        for i in range(len(bins) - 1):
            if bins[i] <= count < bins[i + 1]:
                bin_counts[i] += 1
                break

    # Add a new sheet for word count distribution
    distribution_sheet = result_workbook.create_sheet(title="Word Count Distribution")

    # Insert word count distribution into the new sheet
    distribution_sheet.append(["Range", "Count"])
    for label, count in zip(bin_labels, bin_counts):
        distribution_sheet.append([label, count])

    # Create a bar chart
    chart = BarChart()
    chart.type = 'col'
    chart.title = "Distribution of Word Counts in Chunk Texts"
    chart.x_axis.title = "Word Count Range"
    chart.y_axis.title = "Frequency"

    # Define the data for the chart
    data = Reference(distribution_sheet, min_col=2, min_row=1, max_row=len(bin_counts) + 1, max_col=2)
    categories = Reference(distribution_sheet, min_col=1, min_row=2, max_row=len(bin_counts) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    # Set chart style
    chart.style = 10
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.y_axis.majorGridlines = None
    chart.width = 25
    chart.height = 15
    # Add the chart to the sheet
    distribution_sheet.add_chart(chart, "E5")

    # ************************************************************** #
    # Create a new sheet for chunk type distribution
    chunk_type_sheet = result_workbook.create_sheet(title="Chunk Type Distribution")

    # Count the number of chunks of each type
    chunk_type_counts = {}
    for chunk_type in chunk_types_less_than_20:
        if chunk_type not in chunk_type_counts:
            chunk_type_counts[chunk_type] = [0, 0]    
        chunk_type_counts[chunk_type][0] += 1
    
    for chunk_type in chunk_types_all:
        if chunk_type not in chunk_type_counts:
            chunk_type_counts[chunk_type] = [0, 0]
        chunk_type_counts[chunk_type][1] += 1
    
    # Insert chunk type distribution into the new sheet
    chunk_type_sheet.append(["Chunk Type", "count < 20", "Total Count"])
    for chunk_type, count in chunk_type_counts.items():
        chunk_type_sheet.append([chunk_type, count[0], count[1]])
    
    # Create a bar chart
    chart1 = BarChart()
    chart1.type = 'col'
    chart1.title = "Distribution of Chunk Types for Chunk Length < 20"
    chart1.x_axis.title = "Chunk Type"
    chart1.y_axis.title = "Frequency"

    # Define the data for the chart1
    data = Reference(chunk_type_sheet, min_col=2, min_row=1, max_row=len(chunk_type_counts) + 1, max_col=2)
    categories = Reference(chunk_type_sheet, min_col=1, min_row=2, max_row=len(chunk_type_counts) + 1)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(categories)

    # Set chart1 style
    chart1.style = 10
    chart1.dataLabels = DataLabelList()
    chart1.dataLabels.showVal = True
    chart1.y_axis.majorGridlines = None
    chart1.width = 25
    chart1.height = 17
    # Add the chart1 to the sheet
    chunk_type_sheet.add_chart(chart1, "E5")

    chart2 = BarChart()
    chart2.type = 'col'
    chart2.title = "Distribution of Chunk Types"
    chart2.x_axis.title = "Chunk Type"
    chart2.y_axis.title = "Frequency"

    # Define the data for the chart2
    data2 = Reference(chunk_type_sheet, min_col=3, min_row=1, max_row=len(chunk_type_counts) + 1, max_col=3)
    categories2 = Reference(chunk_type_sheet, min_col=1, min_row=2, max_row=len(chunk_type_counts) + 1)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(categories2)

    # Set chart2 style
    chart2.style = 10
    chart2.dataLabels = DataLabelList()
    chart2.dataLabels.showVal = True
    chart2.y_axis.majorGridlines = None
    chart2.width = 25
    chart2.height = 17
    # Add the chart2 to the sheet
    chunk_type_sheet.add_chart(chart2, "S5")
    # ************************************************************** #
    # Create a new sheet for average word count by chunk type
    avg_word_count_sheet = result_workbook.create_sheet(title="Avg Word Count by Chunk Type")

    # Calculate average word count by chunk type
    chunk_type_word_counts = {}
    for chunk_type, chunk_length in zip(chunk_types_all, chunk_lengths):
        if chunk_type not in chunk_type_word_counts:
            chunk_type_word_counts[chunk_type] = []
        chunk_type_word_counts[chunk_type].append(chunk_length)

    avg_word_counts = {chunk_type: sum(lengths) / len(lengths) for chunk_type, lengths in chunk_type_word_counts.items()}

    # Insert average word count data into the new sheet
    avg_word_count_sheet.append(["Chunk Type", "Average Word Count"])
    for chunk_type, avg_word_count in avg_word_counts.items():
        avg_word_count_sheet.append([chunk_type, avg_word_count])

    # Create a bar chart for average word count by chunk type
    avg_word_count_chart = BarChart()
    avg_word_count_chart.type = 'col'
    avg_word_count_chart.title = "Average Word Count by Chunk Type"
    avg_word_count_chart.x_axis.title = "Chunk Type"
    avg_word_count_chart.y_axis.title = "Average Word Count"

    # Define the data for the chart
    data = Reference(avg_word_count_sheet, min_col=2, min_row=1, max_row=len(avg_word_counts) + 1)
    categories = Reference(avg_word_count_sheet, min_col=1, min_row=2, max_row=len(avg_word_counts) + 1)
    avg_word_count_chart.add_data(data, titles_from_data=True)
    avg_word_count_chart.set_categories(categories)

    # Set chart style
    avg_word_count_chart.style = 10
    avg_word_count_chart.dataLabels = DataLabelList()
    avg_word_count_chart.dataLabels.showVal = True
    avg_word_count_chart.y_axis.majorGridlines = None
    avg_word_count_chart.width = 25
    avg_word_count_chart.height = 15
    # Add the chart to the sheet
    avg_word_count_sheet.add_chart(avg_word_count_chart, "E5")
    # ************************************************************** #
    # Create a new sheet for chunk length vs. query relevance
    relevance_sheet = result_workbook.create_sheet(title="Chunk Length vs. Chunk Score")

    # Insert chunk length and relevance data into the new sheet
    length_bins = [0, 10, 20, 50, 100, 200, 500, 1000]
    binned_relevance = defaultdict(list)
    for length, score in zip(chunk_lengths, chunk_scores):
        for i in range(len(length_bins) - 1):
            if length_bins[i] <= length < length_bins[i + 1]:
                binned_relevance[length_bins[i]].append(score)
                break
    # Calculate average relevance for each bin
    avg_relevance = {bin_start: np.mean(relevances) for bin_start, relevances in binned_relevance.items()}

    relevance_sheet.append(["Chunk Length Range", "Average Score", "Count"])
    for bin_start, avg_rel in sorted(avg_relevance.items()):
        relevance_sheet.append([f"{bin_start}-{length_bins[length_bins.index(bin_start) + 1]}", avg_rel, len(binned_relevance[bin_start])])


    # Create a scatter chart for chunk length vs. query relevance
    relevance_chart = ScatterChart()
    relevance_chart.title = "Chunk Length vs. Chunk Score"
    relevance_chart.x_axis.title = "Chunk Length"
    relevance_chart.y_axis.title = "Chunk Score"

    # Define the data for the chart
    xvalues = Reference(relevance_sheet, min_col=1, min_row=2, max_row=len(chunk_lengths) + 1)
    yvalues = Reference(relevance_sheet, min_col=2, min_row=2, max_row=len(chunk_lengths) + 1)
    series = Series(yvalues, xvalues, title="Score")
    relevance_chart.series.append(series)

    # Set chart style
    relevance_chart.style = 10
    relevance_chart.y_axis.majorGridlines = None
    relevance_chart.width = 25
    relevance_chart.height = 15
    # Add the chart to the sheet
    relevance_sheet.add_chart(relevance_chart, "E5")

    # Save the new Excel file with the results
    result_file_path = './query_results.xlsx'
    result_workbook.save(filename=result_file_path)
    query_workbook.close()
    result_workbook.close()